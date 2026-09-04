import openpyxl
import re
import sqlite3
import json

wb = openpyxl.load_workbook(r'public/storage/PRECIOS EN LINEA SM .xlsx', data_only=True)

excel_items = []

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    curr_model = ''
    for r in range(1, sheet.max_row + 1):
        c1 = sheet.cell(row=r, column=1).value
        c2 = sheet.cell(row=r, column=2).value
        c3 = sheet.cell(row=r, column=3).value
        c4 = sheet.cell(row=r, column=4).value
        
        if c1 and str(c1).strip() != '':
            curr_model = str(c1).strip()
            
        price = None
        if isinstance(c4, (int, float)) and c4 > 0:
            price = float(c4)
        elif isinstance(c3, (int, float)) and c3 > 0 and sheetname == 'Hoja2':
            price = float(c3)
            
        if price:
            type_mueble = str(c2).strip() if c2 else ''
            detail_desc = str(c3).strip() if c3 else ''
            full_text = f"{curr_model} {type_mueble} {detail_desc}".strip()
            excel_items.append({
                'model': curr_model,
                'type': type_mueble,
                'detail': detail_desc,
                'full_text': full_text,
                'price': price,
                'sheet': sheetname,
                'row': r
            })

print(f"Total excel price items found: {len(excel_items)}")
for item in excel_items:
    print(f"[{item['sheet']}:R{item['row']}] Model: {item['model']} | Desc: {item['full_text']} => ${item['price']}")
