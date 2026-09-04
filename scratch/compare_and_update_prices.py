import openpyxl
import re
import sqlite3
import unicodedata

def normalize_text(text):
    if not text:
        return ""
    text = str(text).upper()
    text = unicodedata.normalize('NFD', text).encode('ascii', 'ignore').decode('utf-8')
    text = re.sub(r'[^A-Z0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

# 1. Parse Excel Prices
wb = openpyxl.load_workbook(r'public/storage/PRECIOS EN LINEA SM .xlsx', data_only=True)

excel_price_map = []

# Hoja 1 Parsing
sheet1 = wb['Hoja1']
curr_model = ''
for r in range(3, 84):
    c1 = sheet1.cell(row=r, column=1).value
    c2 = sheet1.cell(row=r, column=2).value
    c3 = sheet1.cell(row=r, column=3).value
    c4 = sheet1.cell(row=r, column=4).value
    
    if c1 and str(c1).strip() != '':
        curr_model = str(c1).strip()
        
    price = None
    if isinstance(c4, (int, float)) and c4 > 0:
        price = float(c4)
        
    if price and curr_model:
        t_mueble = str(c2).strip() if c2 else ''
        t_detail = str(c3).strip() if c3 else ''
        full_raw = f"{curr_model} {t_mueble} {t_detail}".strip()
        excel_price_map.append({
            'model': curr_model,
            'norm_model': normalize_text(curr_model),
            'norm_full': normalize_text(full_raw),
            'price': price,
            'source': f'Hoja1:R{r}'
        })

# Hoja 2 Parsing (Sillas / Salas especiales)
sheet2 = wb['Hoja2']
curr_model = ''
for r in range(1, sheet2.max_row + 1):
    c1 = sheet2.cell(row=r, column=1).value
    c2 = sheet2.cell(row=r, column=2).value
    c3 = sheet2.cell(row=r, column=3).value
    if c1 and str(c1).strip() != '':
        curr_model = str(c1).strip()
    if isinstance(c3, (int, float)) and c3 > 0:
        full_raw = f"{curr_model} {c2 or ''}".strip()
        excel_price_map.append({
            'model': curr_model,
            'norm_model': normalize_text(curr_model),
            'norm_full': normalize_text(full_raw),
            'price': float(c3),
            'source': f'Hoja2:R{r}'
        })

print(f"Total excel prices indexed: {len(excel_price_map)}")

# 2. Connect to MySQL / Laravel Database via PDO or direct query script in PHP
