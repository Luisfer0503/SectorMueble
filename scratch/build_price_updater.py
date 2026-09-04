import openpyxl
import re
import json

# Load Excel
wb = openpyxl.load_workbook(r'public/storage/PRECIOS EN LINEA SM .xlsx', data_only=True)

excel_rows = []

def clean(s):
    if not s: return ""
    s = str(s).upper()
    # Normalize accents
    for orig, rep in [('Á','A'),('É','E'),('Í','I'),('Ó','O'),('Ú','U'),('Ñ','N'),('Ü','U')]:
        s = s.replace(orig, rep)
    s = re.sub(r'[^A-Z0-9\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

# Parse Hoja1
sheet1 = wb['Hoja1']
curr_model = ''
for r in range(3, sheet1.max_row + 1):
    c1 = sheet1.cell(row=r, column=1).value
    c2 = sheet1.cell(row=r, column=2).value
    c3 = sheet1.cell(row=r, column=3).value
    c4 = sheet1.cell(row=r, column=4).value
    
    if c1 and str(c1).strip():
        curr_model = str(c1).strip()
        
    price = None
    if isinstance(c4, (int, float)) and c4 > 0:
        price = float(c4)
        
    if price and curr_model:
        excel_rows.append({
            'model': curr_model,
            'clean_model': clean(curr_model),
            'type': str(c2 or '').strip(),
            'detail': str(c3 or '').strip(),
            'clean_full': clean(f"{curr_model} {c2 or ''} {c3 or ''}"),
            'price': price,
            'sheet': 'Hoja1',
            'row': r
        })

# Parse Hoja2
sheet2 = wb['Hoja2']
curr_model = ''
for r in range(1, sheet2.max_row + 1):
    c1 = sheet2.cell(row=r, column=1).value
    c2 = sheet2.cell(row=r, column=2).value
    c3 = sheet2.cell(row=r, column=3).value
    
    if c1 and str(c1).strip():
        curr_model = str(c1).strip()
        
    price = None
    if isinstance(c3, (int, float)) and c3 > 0:
        price = float(c3)
        
    if price and curr_model:
        excel_rows.append({
            'model': curr_model,
            'clean_model': clean(curr_model),
            'type': str(c2 or '').strip(),
            'detail': '',
            'clean_full': clean(f"{curr_model} {c2 or ''}"),
            'price': price,
            'sheet': 'Hoja2',
            'row': r
        })

with open(r'scratch/excel_prices.json', 'w', encoding='utf-8') as f:
    json.dump(excel_rows, f, ensure_ascii=False, indent=2)

print(f"Exported {len(excel_rows)} excel price entries to scratch/excel_prices.json.")
